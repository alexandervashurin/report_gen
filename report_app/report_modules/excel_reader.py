#!/usr/bin/python3
from openpyxl import load_workbook, Workbook
import datetime
from more_itertools import chunked
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

dt = datetime.datetime.now().date()

sl = ['СТПР', 'Причина отстановки электровоза, электропоезда', 'ПЧ ТК', 'ПЧ ВО', 'БЦВ', 'ПСН', 'УЗАРД',
      'шкаф', 'ПЧТК', 'ПЧВО', 'СТПР1000', 'ПЧ-ТК', 'ПЧ-ВО', 'ПЧЗУ', 'ПЧ ЗУ', 'ПЧ-ЗУ', 'СТПР 1000']


def method_excel(strr: str):
    wb2 = load_workbook(strr)
    res = []
    for row in wb2.active.iter_rows(min_row=1, max_col=8, max_row=2000):
        for cell in row:
            res.append(cell)

    sor = [i.value for i in res]
    sorter = list(chunked(sor[1:], 8))

    unsort = sorter
    s = []
    for ii in unsort:
        if ii[4] is not None:
            for i in sl:
                if i in ii[4]:
                    s.append(ii)
        else:
            if ii[0] is not None:
                s.append(ii)

    wb = Workbook()
    ws = wb.active
    for i in s:
        ws.append(i)

    list_coor_b = [i[1].coordinate for i in list(wb.active) if i[1].value is None]  # '' coordinate B
    list_merged_cells = ["A" + i[1:] + ":" + "G" + i[1:] for i in list_coor_b]
    for i in list_merged_cells:
        ws.merge_cells(i)

    file = '/' + str(dt).split('-')[0] + '/' + str(dt).split('-')[1] + '/' + str(dt) + '.xlsx'
    wb.save(file)

    return sor
